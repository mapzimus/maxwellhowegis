"""
Add the REVENUE / equity side of school finance to the atlas (Session S11).

All 15 existing Finance metrics are expenditure-side (per-pupil spending) or
spending-vs-target ratios. This adds the "who pays" story that the Student
Opportunity Act is about: how much Chapter 70 STATE AID each district gets per
pupil, the REQUIRED LOCAL CONTRIBUTION its member municipalities must raise, and
the share of the foundation budget funded locally vs by the state.

  chapter70_per_pupil          Chapter 70 state aid / foundation enrollment   (USD/pupil)
  required_local_contribution  Required local contribution                    (USD, raw)
  local_share_pct              Required local contribution / foundation budget (0-1 fraction;
                               its complement, 1 - local_share, is the state's share)

Plus the raw underlying dollars used to derive the above (handy in tooltips and
for parity with the existing foundation_budget / required_nss columns):

  ch70_aid                     Chapter 70 aid                                 (USD, raw)
  foundation_enrollment        Foundation (formula) enrollment                (pupil count)

── SOURCE (important: NOT Socrata) ─────────────────────────────────────────────
Chapter 70 state aid and required local contribution are NOT published as a
queryable Socrata table. The Education-to-Career hub's "Chapter 70 Program
Information and Data" catalog entry [qt58-634r] is a type='href' link, confirmed
non-tabular ("no row or column access to non-tabular tables"), and the only
tabular finance dataset [5izv-jyrd] carries foundation budget / required &
actual NSS only — no aid, no local-contribution, no valuation columns. So this
fetcher downloads DESE's published Chapter 70 SUMMARY CHART workbook and parses
it with openpyxl:

  https://www.doe.mass.edu/finance/chapter70/fy2026/chapter-2026-local.xlsx
  sheet 'alldistricts', one row per LEA (operating municipals AND regional
  district totals; the companion -reg.xlsx splits regionals by member town but
  its 999/Total rows carry identical figures, so the -local workbook alone is a
  complete per-district table).

Columns (header row located dynamically by the "LEA" cell):
  LEA | District | Operating status* | Foundation enrollment | Foundation budget
      | Required contribution | Chapter 70 aid | Required net school spending

LEA -> atlas DIST_CODE: zero-pad the LEA to 4 digits and append '0000'
(Abington LEA 1 -> '00010000'; Acton-Boxborough regional LEA 600 -> '06000000').
Every one of the atlas's 281 academic DIST_CODEs ends in '0000', so this is exact.

Non-operating districts (operating status 0 — e.g. Gosnold, which tuitions all
pupils out and runs no schools) have all-zero rows; those become **null, never 0**,
per AGENTS.md, and are dropped. Coverage: 280/281 academic districts (FY2026).

── NOT shipped (sourced search documented, could not source) ────────────────────
  eqv_per_pupil  Equalized property value (EQV) per pupil. EQV is published by
                 the MA Dept. of Local Services (DLS), a DIFFERENT source/domain;
                 the Education-to-Career Socrata catalog returns ZERO results for
                 "equalized valuation" / "EQV" / "property wealth". Pulling it
                 would mean wiring a second, unrelated DLS source (biennial EQV
                 tables). Deferred rather than fabricated. NOTE: local_share_pct
                 already captures the wealth/equity signal EQV-per-pupil was meant
                 to proxy (the Ch70 formula derives the required local
                 contribution from each town's property wealth AND income).

Fiscal year: FY2026 (DESE's latest final Chapter 70 distribution; fresher than
the SY2022 NSS columns the atlas already ships). All dollars kept RAW; the one
ratio (local_share_pct) is a 0-1 fraction.

Output: ``data/ma_district_finance_revenue.json`` :: { DIST_CODE: {col: value} }

Run from repo root::  python scripts/fetch_finance_revenue.py
"""
from __future__ import annotations
import json, urllib.request
from pathlib import Path
from io import BytesIO

import openpyxl

REPO = Path(__file__).resolve().parent.parent
DISTS = REPO / "data" / "ma_academic_districts.geojson"
OUT = REPO / "data" / "ma_district_finance_revenue.json"
FY = 2026
URL = (f"https://www.doe.mass.edu/finance/chapter70/fy{FY}/"
       f"chapter-{FY}-local.xlsx")
UA = {"User-Agent": "ma-education-atlas/1.0 (github.com/mapzimus/ma-education-atlas)"}

# Header labels we need (matched case-insensitively, ignoring embedded newlines
# DESE puts inside some header cells, e.g. "Chapter 70\naid").
WANT = {
    "lea":   ("lea",),
    "stat":  ("operating status",),
    "fenr":  ("foundation enrollment",),
    "fbud":  ("foundation budget",),
    "reqc":  ("required contribution",),
    "c70":   ("chapter 70 aid",),
    "rnss":  ("required net school",),
}


def norm_code(lea) -> str | None:
    """LEA -> zero-padded 8-char atlas DIST_CODE ('0001' + '0000')."""
    try:
        return f"{int(lea):04d}0000"
    except (ValueError, TypeError):
        return None


def to_num(v):
    """Float or None. Treat blanks/non-numeric and exact 0 as MISSING (a 0 here
    means a non-operating district, which must be null — never a 0 that would
    poison choropleths/ranks; see AGENTS.md / data_anomalies.md)."""
    if v is None or v == "":
        return None
    try:
        f = float(v)
    except (ValueError, TypeError):
        return None
    return None if f == 0 else f


def clean_label(s) -> str:
    return " ".join(str(s).split()).lower() if s is not None else ""


def find_columns(ws):
    """Locate the header row (the one containing a 'LEA' cell) and return
    {field: 0-based col index} for every WANT field, plus the header row number."""
    for r in range(1, 16):
        labels = [clean_label(c.value) for c in ws[r]]
        if any(lbl == "lea" for lbl in labels):
            idx = {}
            for field, needles in WANT.items():
                for ci, lbl in enumerate(labels):
                    if any(n in lbl for n in needles):
                        idx[field] = ci
                        break
            missing = [f for f in WANT if f not in idx]
            if missing:
                raise SystemExit(f"header row {r} missing columns: {missing}\n"
                                 f"  found labels: {[l for l in labels if l]}")
            return idx, r
    raise SystemExit("could not find a header row containing 'LEA'")


def main() -> int:
    ours = {f["properties"]["DIST_CODE"]
            for f in json.loads(DISTS.read_text())["features"]}

    print(f"downloading {URL}")
    req = urllib.request.Request(URL, headers=UA)
    with urllib.request.urlopen(req, timeout=120) as r:
        blob = r.read()
    wb = openpyxl.load_workbook(BytesIO(blob), read_only=True, data_only=True)
    ws = wb["alldistricts"] if "alldistricts" in wb.sheetnames else wb[wb.sheetnames[0]]
    idx, hdr = find_columns(ws)

    out: dict[str, dict] = {}
    seen_codes = set()
    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        lea = row[idx["lea"]]
        dc = norm_code(lea)
        if dc is None or dc not in ours or dc in seen_codes:
            continue
        fbud = to_num(row[idx["fbud"]])
        c70 = to_num(row[idx["c70"]])
        reqc = to_num(row[idx["reqc"]])
        fenr = to_num(row[idx["fenr"]])
        rnss = to_num(row[idx["rnss"]])
        # A district with no foundation budget / no aid is non-operating: skip it
        # entirely (don't emit zero-or-null noise for it).
        if not fbud or not c70:
            continue
        seen_codes.add(dc)
        rec: dict[str, float] = {}
        rec["ch70_aid"] = round(c70, 2)
        if reqc is not None:
            rec["required_local_contribution"] = round(reqc, 2)
            rec["local_share_pct"] = round(reqc / fbud, 4)  # 0-1; state share = 1 - this
        if fenr:
            rec["foundation_enrollment"] = round(fenr, 1)
            rec["chapter70_per_pupil"] = round(c70 / fenr, 2)
        out[dc] = rec
    wb.close()

    OUT.write_text(json.dumps(out, indent=1))

    def cov(col):
        return sum(1 for v in out.values() if col in v)

    print(f"wrote {OUT.relative_to(REPO)} for {len(out)} districts "
          f"(of {len(ours)} academic) — FY{FY}, source: DESE Ch.70 summary chart")
    print(f"  chapter70_per_pupil:          {cov('chapter70_per_pupil')} (USD/pupil)")
    print(f"  required_local_contribution:  {cov('required_local_contribution')} (USD)")
    print(f"  local_share_pct:              {cov('local_share_pct')} (0-1 fraction)")
    print(f"  ch70_aid:                     {cov('ch70_aid')} (USD, raw)")
    print(f"  foundation_enrollment:        {cov('foundation_enrollment')} (count)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
